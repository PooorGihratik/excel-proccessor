from .payment_file_parser import PaymentFileParser
from .payment_workbook_register import PaymentWorkbookRegister
from .common import PaymentSection
from .payment_sheet_generator import PaymentSheetGenerator
from .ui_wrappers import UIExcelWrapper, UIWrapable
from openpyxl import Workbook, load_workbook
from dataclasses import dataclass


@dataclass
class SheetRegisterOption:
    account: str
    target_sheet_index: int
    payments_sheet_index: int
    dept_sheet_index: int


@dataclass
class ProcessorOptions:
    file_extension: str
    overall_file_name: str
    output_file_name: str
    overall_output_file: str
    sheets: list[SheetRegisterOption]


class ExcelPaymentProcessor(UIWrapable):
    _config = ProcessorOptions(
        output_file_name='Расшифровка банка',
        overall_file_name='Лицевые счета',
        overall_output_file="Лицевые счета - вывод",
        file_extension='.xlsx',
        sheets=[
            # Utility
            SheetRegisterOption(
                account="40703810211180030006",
                target_sheet_index=0,
                payments_sheet_index=1,
                dept_sheet_index=3
            ),
            # Overhaul
            SheetRegisterOption(
                account="40705810011000000589",
                target_sheet_index=4,
                payments_sheet_index=5,
                dept_sheet_index=7
            ),
        ],
    )

    def __init__(self, config: ProcessorOptions = None):
        super().__init__(UIExcelWrapper)
        if config is not None:
            self.load_options(config)

    def load_options(self, config: ProcessorOptions):
        self._config = config
    
    _overall_read_workbook: Workbook
    _overall_write_workbook: Workbook
    _generated_workbook: Workbook
    _payment_sheet_generator = PaymentSheetGenerator()
    _registerer_dict: dict[str, PaymentWorkbookRegister] = {}
    _count = 0

    @property
    def is_workbooks_loaded(self):
        return self._overall_read_workbook and self._overall_write_workbook and self._generated_workbook

    @property
    def overall_book_name(self):
        return self._config.overall_file_name + self._config.file_extension

    @property
    def overall_output_book_name(self):
        return self._config.overall_output_file + self._config.file_extension

    @property
    def output_book_name(self):
        return self._config.output_file_name + self._config.file_extension

    def load_workbooks(self):
        with self._ui_wrapper.loading_overall(self.overall_book_name):
            self._overall_read_workbook = load_workbook(
                filename=self.overall_book_name,
                data_only=True,
                read_only=True
            )

            self._overall_write_workbook = load_workbook(
                filename=self.overall_book_name
            )

            for sheet_params in self._config.sheets:
                self._registerer_dict[sheet_params.account] = PaymentWorkbookRegister(
                    self._overall_read_workbook.worksheets[sheet_params.target_sheet_index],
                    self._overall_write_workbook.worksheets[sheet_params.payments_sheet_index],
                    self._overall_write_workbook.worksheets[sheet_params.dept_sheet_index]
                )

        with self._ui_wrapper.create_output_workbook(self.output_book_name):
            self._generated_workbook = Workbook()
            self._generated_workbook.remove(self._generated_workbook.worksheets[0])
            self._payment_sheet_generator.workbook = self._generated_workbook

    def close_workbooks(self):
        self._overall_read_workbook.close()
        self._overall_write_workbook.close()
        self._generated_workbook.close()

        self._payment_sheet_generator.workbook = None
        self._registerer_dict = {}

    def save_workbooks(self):
        with self._ui_wrapper.save_book(self.output_book_name):
            self._generated_workbook.save(filename=self.output_book_name)
        with self._ui_wrapper.save_book(self.overall_output_book_name):
            self._overall_write_workbook.save(filename=self.overall_output_book_name)

    def register_sections(self, payments_sections: list[PaymentSection]):
        if not self.is_workbooks_loaded:
            self.load_workbooks()

        for section in payments_sections:
            registerer = self._registerer_dict[section.account]
            if registerer:
                for payment in section.payment_items:
                    # TODO: make an dept direction
                    if payment.period.year == 22: 
                        registerer.add_payment(payment)
                    self._ui_wrapper.payment_process_advance()
            self._payment_sheet_generator.add_payment_section(section)

    def process_files(self, directory: str):
        self._ui_wrapper.status_load_workbooks()
        self.load_workbooks()

        self._ui_wrapper.status_load_files()
        file_parser = PaymentFileParser(directory)
        payment_data = file_parser.parse_all_files_in_directory()

        sum_payments = {}
        for key in payment_data.keys():
            sum_payments[key] = 0

        self._ui_wrapper.status_distribute_payments()
        self._ui_wrapper.payment_start_progress(payment_data)
        for item in payment_data.values():
            for section in item:
                for info in section.payment_items:
                    sum_payments[section.account] += info.payment
            self.register_sections(item)

        self._ui_wrapper.final_sum(sum_payments)
        self.save_workbooks()
        self.close_workbooks()
        self._ui_wrapper.done_print()
        
