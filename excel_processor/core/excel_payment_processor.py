from .payment_workbook_register import PaymentSheetRegister
from .common import PaymentSection, PaymentInfo
from .payment_sheet_generator import PaymentSheetGenerator
from openpyxl import Workbook, load_workbook
from dataclasses import dataclass
from typing import Callable


@dataclass
class SheetRegisterOption:
    account: str
    target_sheet_index: int
    payments_sheet_index: int
    dept_sheet_index: int


@dataclass
class ProcessorOptions:
    file_extension: str
    overall_file_name: str
    output_file_name: str
    overall_output_file: str
    sheets: list[SheetRegisterOption]


class PaymentWorkbookRegister:
    def __init__(self, config: ProcessorOptions = None):
        if config is not None:
            self.load_options(config)

    # Config
    # -----------------------------------------
    _config = ProcessorOptions(
        output_file_name='Расшифровка банка',
        overall_file_name='Лицевые счета',
        overall_output_file="Лицевые счета - вывод",
        file_extension='.xlsx',
        sheets=[
            # Utility
            SheetRegisterOption(
                account="40703810211180030006",
                target_sheet_index=0,
                payments_sheet_index=1,
                dept_sheet_index=3
            ),
            # Overhaul
            SheetRegisterOption(
                account="40705810011000000589",
                target_sheet_index=4,
                payments_sheet_index=5,
                dept_sheet_index=7
            ),
        ],
    )

    def load_options(self, config: ProcessorOptions):
        self._config = config

    # Overall workbook
    # -----------------------------------------
    _overall_read_workbook: Workbook
    _overall_write_workbook: Workbook
    _registerer_dict: dict[str, PaymentSheetRegister] = {}

    @property
    def is_overall_loaded(self):
        return self._overall_read_workbook and self._overall_write_workbook

    @property
    def overall_book_name(self):
        return self._config.overall_file_name + self._config.file_extension

    @property
    def overall_output_book_name(self):
        return self._config.overall_output_file + self._config.file_extension

    def load_overall(self):
        self._overall_read_workbook = load_workbook(
            filename=self.overall_book_name,
            data_only=True,
            read_only=True
        )

        self._overall_write_workbook = load_workbook(
            filename=self.overall_book_name
        )

        for sheet_params in self._config.sheets:
            self._registerer_dict[sheet_params.account] = PaymentSheetRegister(
                self._overall_read_workbook.worksheets[sheet_params.target_sheet_index],
                self._overall_write_workbook.worksheets[sheet_params.payments_sheet_index],
                self._overall_write_workbook.worksheets[sheet_params.dept_sheet_index]
            )

    def close_overall(self):
        self._overall_read_workbook.close()
        self._overall_write_workbook.close()
        self._registerer_dict = {}

    def save_overall(self):
        self._overall_write_workbook.save(filename=self.overall_output_book_name)

    # Bank parser workbook
    # -----------------------------------------
    _bank_workbook: Workbook
    _payment_sheet_generator = PaymentSheetGenerator()

    @property
    def bank_workbook_loaded(self):
        return self._bank_workbook

    @property
    def bank_workbook_name(self):
        return self._config.output_file_name + self._config.file_extension

    def create_bank_workbook(self):
        self._bank_workbook = Workbook()
        self._bank_workbook.remove(self._bank_workbook.worksheets[0])
        self._payment_sheet_generator.workbook = self._bank_workbook

    def close_bank_workbook(self):
        self._bank_workbook.close()
        self._payment_sheet_generator.workbook = None

    def save_bank_workbook(self):
        self._bank_workbook.save(filename=self.bank_workbook_name)

    # Common
    # -----------------------------------------
    def register_sections(
            self,
            payments_sections: list[PaymentSection],
            section_callback: Callable[[PaymentSection], None] = None,
            info_callback: Callable[[PaymentInfo], None] = None
    ):
        for section in payments_sections:
            if section_callback:
                section_callback(section)
            for payment in section.payment_items:
                # TODO: make an dept direction
                if payment.period.year == 22 and section.account in self._registerer_dict.keys():
                    self._registerer_dict[section.account].add_payment(payment)
                if info_callback:
                    info_callback(payment)

            self._payment_sheet_generator.add_payment_section(section)

    def distribute_payments(
            self,
            payment_data: dict[str, list[PaymentSection]],
            section_callback: Callable[[PaymentSection], None] = None,
            info_callback: Callable[[PaymentInfo], None] = None
    ) -> dict[str, int]:
        sum_payments = {}
        for key in payment_data.keys():
            sum_payments[key] = 0
        for item in payment_data.values():
            for section in item:
                for info in section.payment_items:
                    sum_payments[section.account] += info.payment
            self.register_sections(item, section_callback, info_callback)

        return sum_payments
