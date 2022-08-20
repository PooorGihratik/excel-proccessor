from openpyxl.worksheet.worksheet import Worksheet, Cell
from openpyxl.styles import PatternFill, Font
from .common import PaymentInfo
import math


class PaymentSheetRegister:
    _target_sheet: Worksheet
    _payment_sheet: Worksheet
    _dept_sheet: Worksheet

    _columns_per_month: int

    def __init__(self,
                 target_sheet: Worksheet,
                 payment_sheet: Worksheet,
                 dept_sheet: Worksheet,
                 columns_per_month=1
                 ):
        self._payment_sheet = payment_sheet
        self._target_sheet = target_sheet
        self._dept_sheet = dept_sheet
        self._columns_per_month = columns_per_month

    @staticmethod
    def colorize_payment(payment: float, account_target: float) -> PatternFill:
        return \
            PatternFill() \
            if type(account_target) is not float and type(account_target) is not int else \
                PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") \
                    if math.trunc(account_target) == math.trunc(payment) else \
                    PatternFill(start_color="fcd5b5", end_color="fcd5b5", fill_type="solid") \
                        if math.trunc(account_target) > math.trunc(payment) else \
                        PatternFill(start_color="558ed5", end_color="558ed5", fill_type="solid")

    def find_target_cell(self, apartment, month, payment) -> (Cell, Cell):
        const_shift = 2
        shift = 1  # Initial value
        max_shift = -(month - 1) * self._columns_per_month + 1
        row_calc = lambda: apartment + 2
        col_calc = lambda: (month - 1) * self._columns_per_month + shift + const_shift

        account_target_cell = self._target_sheet.cell(
            row=row_calc(),
            column=col_calc()
        )

        # Move forward until reaching white space
        while account_target_cell.value and math.trunc(account_target_cell.value) != math.trunc(payment):
            shift += 1
            account_target_cell = self._target_sheet.cell(
                row=row_calc(),
                column=col_calc()
            )

        # If value does not found
        if not account_target_cell.value:
            shift = 1
            account_target_cell = self._target_sheet.cell(
                row=row_calc(),
                column=col_calc()
            )
            # Move backwards until reaching beginning
            while account_target_cell.value is not None and \
                    math.trunc(account_target_cell.value) != math.trunc(payment):
                shift -= 1
                account_target_cell = self._target_sheet.cell(
                    row=row_calc(),
                    column=col_calc()
                )
                if shift == max_shift:
                    shift = 1
                    break

        target_cell = self._payment_sheet.cell(
            row=row_calc(),
            column=col_calc()
        )

        while target_cell.value is not None:
            shift += 1
            target_cell = self._payment_sheet.cell(
                row=row_calc(),
                column=col_calc()
            )

        account_target_cell = self._target_sheet.cell(
            row=row_calc(),
            column=col_calc()
        )
        return target_cell, account_target_cell

    def add_dept_payment(self, payment_info: PaymentInfo):
        # TODO: Make an dept
        pass
    
    count = 0
    
    def add_payment(self, payment_info: PaymentInfo):
        result = self.find_target_cell(
            payment_info.apartment,
            payment_info.period.month,
            payment_info.payment
        )

        target_cell = result[0]
        account_target_cell = result[1]
        if target_cell.value is None:
            self.count += 1
            target_cell.value = payment_info.payment
            target_cell.fill = self.colorize_payment(payment_info.payment, account_target_cell.value)
            target_cell.font = Font(color="FF0000", size=7)
