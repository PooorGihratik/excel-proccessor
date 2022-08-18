from openpyxl import Workbook
from .common import PaymentSection


class PaymentSheetGenerator:
    workbook: Workbook

    def __init__(self, workbook: Workbook = None):
        self.workbook = workbook

    def _validate_workbook(self):
        if not self.workbook:
            raise TypeError("Workbook does not specified")
        else:
            pass

    def _create_sheet(self, sheet_name: str):
        sheet = self.workbook.create_sheet(title=sheet_name)
        sheet.append([
            "Дата",
            "Кв",
            "ФИО",
            "Период",
            "Оплата",
            "На р/с",
            "%"
        ])

    def add_payment_section(self, payment: PaymentSection):
        self._validate_workbook()
        if payment.account not in [item.title for item in self.workbook.worksheets]:
            self._create_sheet(payment.account)
        sheet = self.workbook[payment.account]
        sheet.append([])
        sheet.append(payment.header_items)
        sheet.append([])
        for payment_info in payment.payment_items:
            sheet.append([
                payment_info.date.strftime('%d-%m-%Y'),
                payment_info.apartment,
                payment_info.fullname,
                payment_info.period.initial,
                payment_info.full_payment,
                payment_info.payment,
                payment_info.percent
            ])
